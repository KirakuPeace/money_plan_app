import json
import re
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(
    page_title="マネープラン",
    page_icon="💰",
    layout="wide",
)
# =========================================================
# Excelを使わない入力項目マスター
# まずは基本情報だけでテストする
# =========================================================
INPUT_ITEMS = [
    {
        "order": 1,
        "section": "基本情報",
        "group": "",
        "key": "basic_year",
        "label": "作成年",
        "value": 2026,
        "type": "year",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 2,
        "section": "基本情報",
        "group": "",
        "key": "you_name",
        "label": "あなたの名前",
        "value": "SATOSHI",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 3,
        "section": "基本情報",
        "group": "",
        "key": "you_age",
        "label": "あなたの年齢",
        "value": 62,
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 4,
        "section": "基本情報",
        "group": "",
        "key": "spouse_name",
        "label": "配偶者の名前",
        "value": "TOMO",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 5,
        "section": "基本情報",
        "group": "",
        "key": "spouse_age",
        "label": "配偶者の年齢",
        "value": 57,
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
        {
        "order": 6,
        "section": "子供情報",
        "group": "",
        "key": "child1_name",
        "label": "第1子の名前",
        "value": "MO",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 7,
        "section": "子供情報",
        "group": "",
        "key": "child1_age",
        "label": "第1子の年齢",
        "value": "27",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 8,
        "section": "子供情報",
        "group": "",
        "key": "child2_name",
        "label": "第2子の名前",
        "value": "MI",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 9,
        "section": "子供情報",
        "group": "",
        "key": "child2_age",
        "label": "第2子の年齢",
        "value": "25",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 10,
        "section": "子供情報",
        "group": "",
        "key": "child3_name",
        "label": "第3子の名前",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 11,
        "section": "子供情報",
        "group": "",
        "key": "child3_age",
        "label": "第3子の年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
        {
        "order": 12,
        "section": "金融資産",
        "group": "",
        "key": "cash",
        "label": "現金・普通預金",
        "value": 50,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 13,
        "section": "金融資産",
        "group": "",
        "key": "time_deposit",
        "label": "定期預金",
        "value": 500,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 14,
        "section": "金融資産",
        "group": "",
        "key": "insurance",
        "label": "保険金",
        "value": 100,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 15,
        "section": "金融資産",
        "group": "",
        "key": "securities",
        "label": "有価証券",
        "value": 500,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 16,
        "section": "金融資産",
        "group": "",
        "key": "other_financial_assets",
        "label": "その他金融資産",
        "value": 50,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
        {
        "order": 17,
        "section": "実物資産",
        "group": "",
        "key": "land_house",
        "label": "住宅・土地",
        "value": 3000,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 18,
        "section": "実物資産",
        "group": "",
        "key": "car",
        "label": "自動車",
        "value": 100,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 19,
        "section": "実物資産",
        "group": "",
        "key": "other_fixed_assets",
        "label": "その他資産",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
        {
        "order": 20,
        "section": "負債",
        "group": "",
        "key": "credit_card",
        "label": "クレジットカード",
        "value": 5,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 21,
        "section": "負債",
        "group": "",
        "key": "car_loan",
        "label": "自動車ローン",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 22,
        "section": "負債",
        "group": "",
        "key": "home_loan",
        "label": "住宅ローン",
        "value": 2000,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 23,
        "section": "負債",
        "group": "",
        "key": "other_debts",
        "label": "その他負債",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
        {
        "order": 24,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_1_from_age",
        "label": "何歳から",
        "value": "62",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 25,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_1_to_age",
        "label": "何歳まで",
        "value": "65",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "you_salary_1_from_age",
        "default_offset": "",
        "min_key": "you_salary_1_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 26,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_1_amount",
        "label": "年収",
        "value": 840,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 27,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_1_takehome_rate",
        "label": "控除率",
        "value": 20,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },
        {
        "order": 28,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_2_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 29,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_2_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "you_salary_2_from_age",
        "default_offset": "",
        "min_key": "you_salary_2_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 30,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_2_amount",
        "label": "年収",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 31,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_2_takehome_rate",
        "label": "控除率",
        "value": 20,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },
    {
        "order": 32,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_3_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 33,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_3_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "you_salary_3_from_age",
        "default_offset": "",
        "min_key": "you_salary_3_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 34,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_3_amount",
        "label": "年収",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 35,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_3_takehome_rate",
        "label": "控除率",
        "value": 20,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },
    {
        "order": 36,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_4_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 37,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_4_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "you_salary_4_from_age",
        "default_offset": "",
        "min_key": "you_salary_4_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 38,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_4_amount",
        "label": "年収",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 39,
        "section": "給与・賞与",
        "group": "あなた",
        "key": "you_salary_4_takehome_rate",
        "label": "控除率",
        "value": 20,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },
        {
        "order": 40,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_1_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 41,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_1_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "spouse_salary_1_from_age",
        "default_offset": "",
        "min_key": "spouse_salary_1_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 42,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_1_amount",
        "label": "年収",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 43,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_1_takehome_rate",
        "label": "控除率",
        "value": 20,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },

    {
        "order": 44,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_2_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 45,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_2_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "spouse_salary_2_from_age",
        "default_offset": "",
        "min_key": "spouse_salary_2_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 46,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_2_amount",
        "label": "年収",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 47,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_2_takehome_rate",
        "label": "控除率",
        "value": 20,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },

    {
        "order": 48,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_3_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 49,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_3_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "spouse_salary_3_from_age",
        "default_offset": "",
        "min_key": "spouse_salary_3_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 50,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_3_amount",
        "label": "年収",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 51,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_3_takehome_rate",
        "label": "控除率",
        "value": 20,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },

    {
        "order": 52,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_4_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 53,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_4_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "spouse_salary_4_from_age",
        "default_offset": "",
        "min_key": "spouse_salary_4_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 54,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_4_amount",
        "label": "年収",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 55,
        "section": "給与・賞与",
        "group": "配偶者",
        "key": "spouse_salary_4_takehome_rate",
        "label": "控除率",
        "value": 20,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },
    {
        "order": 56,
        "section": "退職一時金・企業年金",
        "group": "あなた",
        "key": "you_pension_kind",
        "label": "種別",
        "value": "",
        "type": "select",
        "unit": "",
        "options": "一時金,有期,終身",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 57,
        "section": "退職一時金・企業年金",
        "group": "あなた",
        "key": "you_service_years",
        "label": "勤続年数",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 58,
        "section": "退職一時金・企業年金",
        "group": "あなた",
        "key": "you_start_age",
        "label": "受取開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 59,
        "section": "退職一時金・企業年金",
        "group": "あなた",
        "key": "you_period",
        "label": "受取期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 60,
        "section": "退職一時金・企業年金",
        "group": "あなた",
        "key": "you_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 61,
        "section": "退職一時金・企業年金",
        "group": "配偶者",
        "key": "spouse_pension_kind",
        "label": "種別",
        "value": "",
        "type": "select",
        "unit": "",
        "options": "一時金,有期,終身",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 62,
        "section": "退職一時金・企業年金",
        "group": "配偶者",
        "key": "spouse_service_years",
        "label": "勤続年数",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 63,
        "section": "退職一時金・企業年金",
        "group": "配偶者",
        "key": "spouse_start_age",
        "label": "受取開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 64,
        "section": "退職一時金・企業年金",
        "group": "配偶者",
        "key": "spouse_period",
        "label": "受取期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 65,
        "section": "退職一時金・企業年金",
        "group": "配偶者",
        "key": "spouse_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 66,
        "section": "公的年金",
        "group": "あなた",
        "key": "you_basic_pension",
        "label": "基礎年金",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 67,
        "section": "公的年金",
        "group": "あなた",
        "key": "you_basic_pension_start_age",
        "label": "受給開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 68,
        "section": "公的年金",
        "group": "あなた",
        "key": "you_employee_pension",
        "label": "厚生年金",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 69,
        "section": "公的年金",
        "group": "あなた",
        "key": "you_employee_pension_start_age",
        "label": "受給開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 70,
        "section": "公的年金",
        "group": "配偶者",
        "key": "spouse_basic_pension",
        "label": "基礎年金",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 71,
        "section": "公的年金",
        "group": "配偶者",
        "key": "spouse_basic_pension_start_age",
        "label": "受給開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 72,
        "section": "公的年金",
        "group": "配偶者",
        "key": "spouse_employee_pension",
        "label": "厚生年金",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 73,
        "section": "公的年金",
        "group": "配偶者",
        "key": "spouse_employee_pension_start_age",
        "label": "受給開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 74,
        "section": "公的年金",
        "group": "控除率",
        "key": "public_pension_rate",
        "label": "年金控除率",
        "value": 15,
        "type": "percentage",
        "unit": "％",
        "options": "",
        "note": "公的年金・企業年金に対する概算控除率",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": 0,
        "max_value": 100,
    },
    {
        "order": 75,
        "section": "その他の収入",
        "group": "雇用保険",
        "key": "you_unemployment_insurance_age",
        "label": "あなた・受取年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 76,
        "section": "その他の収入",
        "group": "雇用保険",
        "key": "you_unemployment_insurance_amount",
        "label": "あなた・金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 77,
        "section": "その他の収入",
        "group": "雇用保険",
        "key": "spouse_unemployment_insurance_age",
        "label": "配偶者・受取年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 78,
        "section": "その他の収入",
        "group": "雇用保険",
        "key": "spouse_unemployment_insurance_amount",
        "label": "配偶者・金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 79,
        "section": "その他の収入",
        "group": "個人年金",
        "key": "you_personal_pension_start_age",
        "label": "あなた・開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 80,
        "section": "その他の収入",
        "group": "個人年金",
        "key": "you_personal_pension_start_period",
        "label": "あなた・受取期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 81,
        "section": "その他の収入",
        "group": "個人年金",
        "key": "you_personal_pension_start_amount",
        "label": "あなた・金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 82,
        "section": "その他の収入",
        "group": "個人年金",
        "key": "spouse_personal_pension_start_age",
        "label": "配偶者・開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 83,
        "section": "その他の収入",
        "group": "個人年金",
        "key": "spouse_personal_pension_start_period",
        "label": "配偶者・受取期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 84,
        "section": "その他の収入",
        "group": "個人年金",
        "key": "spouse_personal_pension_start_amount",
        "label": "配偶者・金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 85,
        "section": "その他の収入",
        "group": "その他収入①",
        "key": "other_income1_start_age",
        "label": "①・開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 86,
        "section": "その他の収入",
        "group": "その他収入①",
        "key": "other_income1_period",
        "label": "①・期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 87,
        "section": "その他の収入",
        "group": "その他収入①",
        "key": "other_income1_amount",
        "label": "①・金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 88,
        "section": "その他の収入",
        "group": "その他収入②",
        "key": "other_income2_start_age",
        "label": "②・開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 89,
        "section": "その他の収入",
        "group": "その他収入②",
        "key": "other_income2_period",
        "label": "②・期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 90,
        "section": "その他の収入",
        "group": "その他収入②",
        "key": "other_income2_amount",
        "label": "②・金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 91,
        "section": "基礎生活費",
        "group": "基礎生活費①",
        "key": "basic_living_cost1_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 92,
        "section": "基礎生活費",
        "group": "基礎生活費①",
        "key": "basic_living_cost1_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "basic_living_cost1_from_age",
        "default_offset": "",
        "min_key": "basic_living_cost1_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 93,
        "section": "基礎生活費",
        "group": "基礎生活費①",
        "key": "basic_living_cost1_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 94,
        "section": "基礎生活費",
        "group": "基礎生活費②",
        "key": "basic_living_cost2_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 95,
        "section": "基礎生活費",
        "group": "基礎生活費②",
        "key": "basic_living_cost2_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "basic_living_cost2_from_age",
        "default_offset": "",
        "min_key": "basic_living_cost2_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 96,
        "section": "基礎生活費",
        "group": "基礎生活費②",
        "key": "basic_living_cost2_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 97,
        "section": "基礎生活費",
        "group": "基礎生活費③",
        "key": "basic_living_cost3_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 98,
        "section": "基礎生活費",
        "group": "基礎生活費③",
        "key": "basic_living_cost3_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "basic_living_cost3_from_age",
        "default_offset": "",
        "min_key": "basic_living_cost3_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 99,
        "section": "基礎生活費",
        "group": "基礎生活費③",
        "key": "basic_living_cost3_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 100,
        "section": "基礎生活費",
        "group": "基礎生活費④",
        "key": "basic_living_cost4_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 101,
        "section": "基礎生活費",
        "group": "基礎生活費④",
        "key": "basic_living_cost4_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "basic_living_cost4_from_age",
        "default_offset": "",
        "min_key": "basic_living_cost4_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 102,
        "section": "基礎生活費",
        "group": "基礎生活費④",
        "key": "basic_living_cost4_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 103,
        "section": "基礎生活費",
        "group": "基礎生活費⑤",
        "key": "basic_living_cost5_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 104,
        "section": "基礎生活費",
        "group": "基礎生活費⑤",
        "key": "basic_living_cost5_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "basic_living_cost5_from_age",
        "default_offset": "",
        "min_key": "basic_living_cost5_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 105,
        "section": "基礎生活費",
        "group": "基礎生活費⑤",
        "key": "basic_living_cost5_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 106,
        "section": "教育費",
        "group": "教育費①",
        "key": "educational_cost1_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 107,
        "section": "教育費",
        "group": "教育費①",
        "key": "educational_cost1_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "educational_cost1_from_age",
        "default_offset": "",
        "min_key": "educational_cost1_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 108,
        "section": "教育費",
        "group": "教育費①",
        "key": "educational_cost1_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 109,
        "section": "教育費",
        "group": "教育費②",
        "key": "educational_cost2_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 110,
        "section": "教育費",
        "group": "教育費②",
        "key": "educational_cost2_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "educational_cost2_from_age",
        "default_offset": "",
        "min_key": "educational_cost2_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 111,
        "section": "教育費",
        "group": "教育費②",
        "key": "educational_cost2_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 112,
        "section": "教育費",
        "group": "教育費③",
        "key": "educational_cost3_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 113,
        "section": "教育費",
        "group": "教育費③",
        "key": "educational_cost3_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "educational_cost3_from_age",
        "default_offset": "",
        "min_key": "educational_cost3_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 114,
        "section": "教育費",
        "group": "教育費③",
        "key": "educational_cost3_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 115,
        "section": "教育費",
        "group": "教育費④",
        "key": "educational_cost4_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 116,
        "section": "教育費",
        "group": "教育費④",
        "key": "educational_cost4_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "educational_cost4_from_age",
        "default_offset": "",
        "min_key": "educational_cost4_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 117,
        "section": "教育費",
        "group": "教育費④",
        "key": "educational_cost4_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 118,
        "section": "住居費",
        "group": "住居費①",
        "key": "housing_cost1_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 119,
        "section": "住居費",
        "group": "住居費①",
        "key": "housing_cost1_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "housing_cost1_from_age",
        "default_offset": "",
        "min_key": "housing_cost1_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 120,
        "section": "住居費",
        "group": "住居費①",
        "key": "housing_cost1_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 121,
        "section": "住居費",
        "group": "住居費②",
        "key": "housing_cost2_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 122,
        "section": "住居費",
        "group": "住居費②",
        "key": "housing_cost2_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "housing_cost2_from_age",
        "default_offset": "",
        "min_key": "housing_cost2_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 123,
        "section": "住居費",
        "group": "住居費②",
        "key": "housing_cost2_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 124,
        "section": "住居費",
        "group": "住居費③",
        "key": "housing_cost3_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 125,
        "section": "住居費",
        "group": "住居費③",
        "key": "housing_cost3_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "housing_cost3_from_age",
        "default_offset": "",
        "min_key": "housing_cost3_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 126,
        "section": "住居費",
        "group": "住居費③",
        "key": "housing_cost3_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 127,
        "section": "住居費",
        "group": "住居費④",
        "key": "housing_cost4_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 128,
        "section": "住居費",
        "group": "住居費④",
        "key": "housing_cost4_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "housing_cost4_from_age",
        "default_offset": "",
        "min_key": "housing_cost4_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 129,
        "section": "住居費",
        "group": "住居費④",
        "key": "housing_cost4_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 130,
        "section": "保険料",
        "group": "保険料①",
        "key": "insurance1_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 131,
        "section": "保険料",
        "group": "保険料①",
        "key": "insurance1_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "insurance1_from_age",
        "default_offset": "",
        "min_key": "insurance1_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 132,
        "section": "保険料",
        "group": "保険料①",
        "key": "insurance1_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 133,
        "section": "保険料",
        "group": "保険料②",
        "key": "insurance2_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 134,
        "section": "保険料",
        "group": "保険料②",
        "key": "insurance2_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "insurance2_from_age",
        "default_offset": "",
        "min_key": "insurance2_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 135,
        "section": "保険料",
        "group": "保険料②",
        "key": "insurance2_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 136,
        "section": "保険料",
        "group": "保険料③",
        "key": "insurance3_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 137,
        "section": "保険料",
        "group": "保険料③",
        "key": "insurance3_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "insurance3_from_age",
        "default_offset": "",
        "min_key": "insurance3_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 138,
        "section": "保険料",
        "group": "保険料③",
        "key": "insurance3_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 139,
        "section": "保険料",
        "group": "保険料④",
        "key": "insurance4_from_age",
        "label": "何歳から",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 140,
        "section": "保険料",
        "group": "保険料④",
        "key": "insurance4_to_age",
        "label": "何歳まで",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "insurance4_from_age",
        "default_offset": "",
        "min_key": "insurance4_from_age",
        "min_offset": 0,
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 141,
        "section": "保険料",
        "group": "保険料④",
        "key": "insurance4_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 142,
        "section": "その他の支出",
        "group": "その他の支出①",
        "key": "other_expense1_start_age",
        "label": "開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 143,
        "section": "その他の支出",
        "group": "その他の支出①",
        "key": "other_expense1_period",
        "label": "期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 144,
        "section": "その他の支出",
        "group": "その他の支出①",
        "key": "other_expense1_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 145,
        "section": "その他の支出",
        "group": "その他の支出②",
        "key": "other_expense2_start_age",
        "label": "開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 146,
        "section": "その他の支出",
        "group": "その他の支出②",
        "key": "other_expense2_period",
        "label": "期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 147,
        "section": "その他の支出",
        "group": "その他の支出②",
        "key": "other_expense2_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 148,
        "section": "その他の支出",
        "group": "その他の支出③",
        "key": "other_expense3_start_age",
        "label": "開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 149,
        "section": "その他の支出",
        "group": "その他の支出③",
        "key": "other_expense3_period",
        "label": "期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 150,
        "section": "その他の支出",
        "group": "その他の支出③",
        "key": "other_expense3_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 151,
        "section": "その他の支出",
        "group": "その他の支出④",
        "key": "other_expense4_start_age",
        "label": "開始年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 152,
        "section": "その他の支出",
        "group": "その他の支出④",
        "key": "other_expense4_period",
        "label": "期間",
        "value": "",
        "type": "years",
        "unit": "年",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 153,
        "section": "その他の支出",
        "group": "その他の支出④",
        "key": "other_expense4_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 154,
        "section": "その他の支出",
        "group": "一時支出①",
        "key": "primary_expense1_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 155,
        "section": "その他の支出",
        "group": "一時支出①",
        "key": "primary_expense1_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 156,
        "section": "その他の支出",
        "group": "一時支出①",
        "key": "primary_expense1_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 157,
        "section": "その他の支出",
        "group": "一時支出②",
        "key": "primary_expense2_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 158,
        "section": "その他の支出",
        "group": "一時支出②",
        "key": "primary_expense2_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 159,
        "section": "その他の支出",
        "group": "一時支出②",
        "key": "primary_expense2_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 160,
        "section": "その他の支出",
        "group": "一時支出③",
        "key": "primary_expense3_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 161,
        "section": "その他の支出",
        "group": "一時支出③",
        "key": "primary_expense3_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 162,
        "section": "その他の支出",
        "group": "一時支出③",
        "key": "primary_expense3_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 163,
        "section": "その他の支出",
        "group": "一時支出④",
        "key": "primary_expense4_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 164,
        "section": "その他の支出",
        "group": "一時支出④",
        "key": "primary_expense4_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 165,
        "section": "その他の支出",
        "group": "一時支出④",
        "key": "primary_expense4_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 166,
        "section": "その他の支出",
        "group": "一時支出⑤",
        "key": "primary_expense5_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 167,
        "section": "その他の支出",
        "group": "一時支出⑤",
        "key": "primary_expense5_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 168,
        "section": "その他の支出",
        "group": "一時支出⑤",
        "key": "primary_expense5_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 169,
        "section": "その他の支出",
        "group": "一時支出⑥",
        "key": "primary_expense6_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 170,
        "section": "その他の支出",
        "group": "一時支出⑥",
        "key": "primary_expense6_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 171,
        "section": "その他の支出",
        "group": "一時支出⑥",
        "key": "primary_expense6_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 172,
        "section": "その他の支出",
        "group": "一時支出⑦",
        "key": "primary_expense7_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 173,
        "section": "その他の支出",
        "group": "一時支出⑦",
        "key": "primary_expense7_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 174,
        "section": "その他の支出",
        "group": "一時支出⑦",
        "key": "primary_expense7_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 175,
        "section": "その他の支出",
        "group": "一時支出⑧",
        "key": "primary_expense8_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 176,
        "section": "その他の支出",
        "group": "一時支出⑧",
        "key": "primary_expense8_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 177,
        "section": "その他の支出",
        "group": "一時支出⑧",
        "key": "primary_expense8_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 178,
        "section": "その他の支出",
        "group": "一時支出⑨",
        "key": "primary_expense9_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 179,
        "section": "その他の支出",
        "group": "一時支出⑨",
        "key": "primary_expense9_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 180,
        "section": "その他の支出",
        "group": "一時支出⑨",
        "key": "primary_expense9_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 181,
        "section": "その他の支出",
        "group": "一時支出⑩",
        "key": "primary_expense10_name",
        "label": "事由",
        "value": "",
        "type": "text",
        "unit": "",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 182,
        "section": "その他の支出",
        "group": "一時支出⑩",
        "key": "primary_expense10_age",
        "label": "支出年齢",
        "value": "",
        "type": "age",
        "unit": "歳",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
    {
        "order": 183,
        "section": "その他の支出",
        "group": "一時支出⑩",
        "key": "primary_expense10_amount",
        "label": "金額",
        "value": 0,
        "type": "money",
        "unit": "万円",
        "options": "",
        "note": "",
        "enabled_condition": "",
        "default_from_key": "",
        "default_offset": "",
        "min_key": "",
        "min_offset": "",
        "max_key": "",
        "max_offset": "",
        "min_value": "",
        "max_value": "",
    },
]

# =========================================================
# 基本設定
# =========================================================
INPUT_FILE = "mp.xlsx"
SHEET_NAME = "入力データ"


# =========================================================
# 表示用フォーマット
# =========================================================
NUMERIC_TYPES = ("money", "age", "year", "years", "number", "percentage")

def fmt_display(value, ftype):
    """
    Streamlit画面に表示するための整形
    例：500 → 500万円、65 → 65歳、20 → 20％
    """
    if value is None:
        return ""

    text = str(value).strip()
    if text == "":
        return ""

    if ftype in ("money", "age", "year", "years", "number", "percentage"):
        digits = re.sub(r"\D", "", text)

        if digits == "":
            return ""

        number = int(digits)

        if ftype == "money":
            return f"{number:,}万円"
        elif ftype == "age":
            return f"{number}歳"
        elif ftype in ("year", "years"):
            return f"{number}年"
        elif ftype == "percentage":
            return f"{number}％"
        else:
            return str(number)

    return text


# =========================================================
# Excel保存用の値に戻す
# =========================================================
def to_excel_value(value, ftype):
    """
    画面上の「500万円」「65歳」「20％」などを、
    Excel保存用に 500, 65, 20 のような数値に戻す
    """
    if value is None:
        return None

    text = str(value).strip()
    if text == "":
        return None

    if ftype in ("money", "age", "year", "years", "number", "percentage"):
        digits = re.sub(r"\D", "", text)
        return int(digits) if digits else None

    return text

def get_number_by_key(rows, ref_key):
    """
    ref_key で指定された項目の現在値を数値で取得する。
    - ref_key が 100 のような数値なら、そのまま 100 として返す
    - それ以外は value_ → input_ → rows の順に key を探す
    """
    if ref_key is None or str(ref_key).strip() == "":
        return None

    ref_key = str(ref_key).strip()

    # 100 のような固定数値なら、そのまま返す
    if re.fullmatch(r"\d+", ref_key):
        return int(ref_key)

    value_key = f"value_{ref_key}"
    widget_key = f"input_{ref_key}"
   
    # まず保存用 session_state を見る
    if value_key in st.session_state:
        value = st.session_state.get(value_key)
        digits = re.sub(r"\D", "", str(value or ""))
        if digits != "":
            return int(digits)

    # 次に画面上の入力値を見る
    if widget_key in st.session_state:
        value = st.session_state.get(widget_key)
        digits = re.sub(r"\D", "", str(value or ""))
        if digits != "":
            return int(digits)

    # 最後に Excel 由来の rows を見る
    for item in rows:
        if str(item.get("key") or "").strip() == ref_key:
            value = item.get("value")
            digits = re.sub(r"\D", "", str(value or ""))
            return int(digits) if digits else None

    return None

# =========================================================
# no_excel版：現在の入力内容をJSON保存用データにする
# =========================================================
def create_save_data(rows):
    """
    現在の画面入力値を key: value の形で保存する。
    """
    save_data = {}

    for item in rows:
        key = str(item.get("key") or "").strip()

        if key == "":
            continue

        value_key = f"value_{key}"
        widget_key = f"input_{key}"

        if value_key in st.session_state:
            value = st.session_state.get(value_key)
        elif widget_key in st.session_state:
            value = st.session_state.get(widget_key)
        else:
            value = item.get("value")

        save_data[key] = value

    return save_data

# =========================================================
# no_excel版：JSONから読み込んだ値を画面入力に反映する
# =========================================================
def load_save_data_to_session(save_data):
    """
    JSONから読み込んだ key: value のデータを
    st.session_state に反映する。
    """
    if not isinstance(save_data, dict):
        st.warning("読み込んだファイルの形式が正しくありません。")
        return

    for key, value in save_data.items():
        key = str(key).strip()

        if key == "":
            continue

        widget_key = f"input_{key}"
        value_key = f"value_{key}"

        st.session_state[widget_key] = value
        st.session_state[value_key] = value

# =========================================================
# keyから文字列を取得する
# =========================================================
def get_text_by_key(rows, ref_key):
    """
    ref_key で指定された項目の現在値を文字列で取得する。
    value_ → input_ → rows の順に探す。
    """
    if ref_key is None or str(ref_key).strip() == "":
        return ""

    ref_key = str(ref_key).strip()

    widget_key = f"input_{ref_key}"
    value_key = f"value_{ref_key}"

    if value_key in st.session_state:
        return str(st.session_state.get(value_key) or "").strip()

    if widget_key in st.session_state:
        return str(st.session_state.get(widget_key) or "").strip()

    for item in rows:
        if str(item.get("key") or "").strip() == ref_key:
            return str(item.get("value") or "").strip()

    return ""

def apply_input_rules(value, item, rows):
    """
    min_key / min_offset / max_key / max_offset に基づいて入力値を補正する
    """
    ftype = str(item.get("type") or "text").strip()

    if ftype not in NUMERIC_TYPES:
        return value

    digits = re.sub(r"\D", "", str(value or ""))
    if digits == "":
        return value

    number = int(digits)

    # min_key + min_offset
    min_key = item.get("min_key")
    min_by_key = get_number_by_key(rows, min_key)

    min_offset = item.get("min_offset")
    try:
        min_offset = int(min_offset) if min_offset not in (None, "") else 0
    except Exception:
        min_offset = 0

    if min_by_key is not None:
        min_limit = min_by_key + min_offset
        if number < min_limit:
            number = min_limit

    # max_key + max_offset
    max_key = item.get("max_key")
    max_by_key = get_number_by_key(rows, max_key)

    max_offset = item.get("max_offset")
    try:
        max_offset = int(max_offset) if max_offset not in (None, "") else 0
    except Exception:
        max_offset = 0

    if max_by_key is not None:
        max_limit = max_by_key + max_offset
        if number > max_limit:
            number = max_limit

    return number


# =========================================================
# 入力値が変わったときに単位付き表示へ整える
# =========================================================
def on_change_format(widget_key, ftype, item=None, rows=None):
    value = st.session_state.get(widget_key, "")

    if item is not None and rows is not None:
        value = apply_input_rules(value, item, rows)

    formatted_value = fmt_display(value, ftype)

    # 画面表示用
    st.session_state[widget_key] = formatted_value

    # 画面から消えても残す保存用
    if widget_key.startswith("input_"):
        key = widget_key.replace("input_", "", 1)
        value_key = f"value_{key}"
        st.session_state[value_key] = formatted_value


# =========================================================
# 入力データシートを読む
# =========================================================
def load_input_rows():
    file_path = Path(INPUT_FILE)

    if not file_path.exists():
        st.error(f"{INPUT_FILE} が見つかりません。app_input_test.py と同じフォルダに置いてください。")
        st.stop()

    wb = load_workbook(INPUT_FILE)
    if SHEET_NAME not in wb.sheetnames:
        st.error(f"{SHEET_NAME} シートが見つかりません。")
        st.stop()

    ws = wb[SHEET_NAME]

    # 1行目の見出しを取得
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # 見出し名 → 列番号
    col_map = {}
    for idx, name in enumerate(headers, start=1):
        if name:
            col_map[str(name).strip()] = idx

    required_cols = ["order", "section", "key", "label", "value", "type"]
    missing = [c for c in required_cols if c not in col_map]

    if missing:
        st.error(f"入力データシートに必要な列がありません: {missing}")
        st.stop()

    rows = []

    for row_no in range(2, ws.max_row + 1):
        item = {}

        for col_name, col_no in col_map.items():
            item[col_name] = ws.cell(row=row_no, column=col_no).value

        # key が空の行は無視
        if item.get("key") is None or str(item.get("key")).strip() == "":
            continue

        item["_row_no"] = row_no
        rows.append(item)

    # order順に並べる
    rows.sort(key=lambda x: x.get("order") or 999999)

    return rows, col_map


# =========================================================
# enabled_condition を判定する
# 例：you_pension_kind=一時金
# =========================================================
def is_enabled(condition):
    if condition is None or str(condition).strip() == "":
        return True

    condition = str(condition).strip()

    if "=" not in condition:
        return True

    key, expected = condition.split("=", 1)
    key = key.strip()
    expected = expected.strip()

    widget_key = f"input_{key}"
    actual = st.session_state.get(widget_key, "")

    return str(actual).strip() == expected


def is_enabled(condition):
    if condition is None or str(condition).strip() == "":
        return True

    condition = str(condition).strip()

    if "=" not in condition:
        return True

    key, expected = condition.split("=", 1)
    key = key.strip()
    expected = expected.strip()

    widget_key = f"input_{key}"
    actual = st.session_state.get(widget_key, "")

    return str(actual).strip() == expected


# =========================================================
# 1項目分の入力ウィジェットを描画する
# =========================================================
def render_input_item(item, rows):
    key = str(item["key"]).strip()
    widget_key = f"input_{key}"
    value_key = f"value_{key}"

    label = str(item.get("label") or key).strip()
    ftype = str(item.get("type") or "text").strip()
    unit = item.get("unit")
    options = item.get("options")
    note = item.get("note")
    condition = item.get("enabled_condition")

    disabled = not is_enabled(condition)

    help_text = None
    if note is not None and str(note).strip() != "":
        help_text = str(note)

    # 画面から消えていた入力欄を再表示するとき、
    # 保存用 value_... があれば、それを input_... に戻す
    if value_key in st.session_state:
        st.session_state[widget_key] = st.session_state[value_key]

    # 選択式
    if ftype == "select":
        if options is None or str(options).strip() == "":
            st.warning(f"{label} の options が空です。")
            return

        option_list = [x.strip() for x in str(options).split(",") if x.strip()]

        current = st.session_state.get(widget_key, "")
        if current not in option_list:
            current = option_list[0]
            st.session_state[widget_key] = current
            st.session_state[value_key] = current

        st.selectbox(
            label,
            option_list,
            key=widget_key,
            disabled=disabled,
            help=help_text,
            on_change=on_change_format,
            args=(widget_key, ftype, item, rows),
        )

    # 数値系
    elif ftype in ("money", "age", "year", "years", "number", "percentage"):
        placeholder = str(unit) if unit is not None else ""

        st.text_input(
            label,
            key=widget_key,
            placeholder=placeholder,
            disabled=disabled,
            help=help_text,
            on_change=on_change_format,
            args=(widget_key, ftype, item, rows),
        )

    # 文字系
    else:
        st.text_input(
            label,
            key=widget_key,
            disabled=disabled,
            help=help_text,
            on_change=on_change_format,
            args=(widget_key, ftype, item, rows),
        )


# =========================================================
# 指定した key の項目を、6カラムの左側に横並び表示する
# =========================================================
def render_row_by_keys(section_rows, key_list, col_widths=None, all_rows=None):
    item_map = {
        str(item["key"]).strip(): item
        for item in section_rows
    }

    # 幅指定がなければ、6カラム均等
    if col_widths is None:
        col_widths = [1, 1, 1, 1, 1, 1]

    cols = st.columns(col_widths, gap="small")

    for i, key in enumerate(key_list):
        item = item_map.get(key)
        if item is None:
            cols[i].warning(f"{key} がありません")
            continue

        with cols[i]:
            render_input_item(item, all_rows or section_rows)


# =========================================================
# 1行目だけ表示し、2行目以降は表示/非表示で切り替える
# show_more が渡された場合は、関数内ではボタンを出さない
# =========================================================
def render_rows_with_more(
    section_rows,
    key_rows,
    col_widths,
    all_rows=None,
    more_label="MORE",
    show_more=None,
):
    """
    key_rows:
      [
        ["xxx1_from_age", "xxx1_to_age", "xxx1_amount"],
        ["xxx2_from_age", "xxx2_to_age", "xxx2_amount"],
        ...
      ]

    1行目は常時表示。
    2行目以降は表示/非表示で切り替える。

    show_more:
      None の場合      → この関数内で表示/非表示ボタンを出す
      True / False の場合 → 外側で作ったボタンの状態を使う
    """
    if not key_rows:
        return

    # 1行目は常時表示
    render_row_by_keys(
        section_rows,
        key_rows[0],
        col_widths=col_widths,
        all_rows=all_rows,
    )

    # 2行目以降がなければ終了
    if len(key_rows) <= 1:
        return

    # 外側から show_more が渡されていない場合だけ、
    # この関数内で表示/非表示ボタンを出す
    if show_more is None:
        more_key = "show_more_" + "_".join(key_rows[0])

        if more_key not in st.session_state:
            st.session_state[more_key] = False

        button_label = "閉じる ▲" if st.session_state[more_key] else "表示 ▼"

        more_col, blank_col = st.columns([1, 2])

        with more_col:
            if st.button(button_label, key=more_key + "_button"):
                st.session_state[more_key] = not st.session_state[more_key]
                st.rerun()

        show_more = st.session_state[more_key]

    # 表示状態なら2行目以降を表示
    if show_more:
        for key_list in key_rows[1:]:
            render_row_by_keys(
                section_rows,
                key_list,
                col_widths=col_widths,
                all_rows=all_rows,
            )


def make_period_key_rows(base_key, periods, suffixes):
    """
    例：
      base_key = "basic_living_cost"
      periods = 5
      suffixes = ["from_age", "to_age", "amount"]

    ↓

      [
        ["basic_living_cost1_from_age", "basic_living_cost1_to_age", "basic_living_cost1_amount"],
        ["basic_living_cost2_from_age", "basic_living_cost2_to_age", "basic_living_cost2_amount"],
        ...
      ]
    """
    key_rows = []

    for i in range(1, periods + 1):
        key_rows.append([
            f"{base_key}{i}_{suffix}"
            for suffix in suffixes
        ])

    return key_rows

# =========================================================
# 給与・賞与を年齢に応じて取得する
# 控除率を反映した手取り額を返す
# =========================================================
def get_salary_income(rows, prefix, age, periods=4):
    """
    prefix:
      あなた   → "you"
      配偶者   → "spouse"

    age:
      判定する年齢

    戻り値：
      指定年齢に該当する給与・賞与の手取り額（万円）
    """
    total = 0

    for i in range(1, periods + 1):
        from_age = get_number_by_key(rows, f"{prefix}_salary_{i}_from_age")
        to_age = get_number_by_key(rows, f"{prefix}_salary_{i}_to_age")
        amount = get_number_by_key(rows, f"{prefix}_salary_{i}_amount")
        takehome_rate = get_number_by_key(rows, f"{prefix}_salary_{i}_takehome_rate")

        if from_age is None or to_age is None or amount is None:
            continue

        # 控除率が未入力なら 0％ として扱う
        if takehome_rate is None:
            takehome_rate = 0

        if from_age <= age <= to_age:
            net_amount = round(amount * (100 - takehome_rate) / 100)
            total += net_amount

    return total

# =========================================================
# 退職所得控除額を計算する
# 単位：万円
# =========================================================
def calc_retirement_income_deduction(service_years):
    """
    service_years:
      勤続年数

    戻り値：
      退職所得控除額（万円）
    """
    if service_years is None:
        return 0

    if service_years <= 20:
        deduction = 40 * service_years
        return max(deduction, 80)

    return 800 + 70 * (service_years - 20)


# =========================================================
# 退職所得にかかる所得税・復興特別所得税を計算する
# 単位：万円
# =========================================================
def calc_retirement_income_tax(taxable_retirement_income):
    """
    taxable_retirement_income:
      課税退職所得金額（万円）

    戻り値：
      所得税・復興特別所得税（万円）
    """
    x = taxable_retirement_income

    if x <= 0:
        return 0

    if x <= 195:
        tax = x * 0.05
    elif x <= 330:
        tax = x * 0.10 - 9.75
    elif x <= 695:
        tax = x * 0.20 - 42.75
    elif x <= 900:
        tax = x * 0.23 - 63.6
    elif x <= 1800:
        tax = x * 0.33 - 153.6
    elif x <= 4000:
        tax = x * 0.40 - 279.6
    else:
        tax = x * 0.45 - 479.6

    # 復興特別所得税を含める
    return round(tax * 1.021)


# =========================================================
# 退職一時金の税額・手取り額を計算する
# 単位：万円
# =========================================================
def calc_retirement_lump_sum_net(amount, service_years):
    """
    amount:
      退職一時金の額面（万円）

    service_years:
      勤続年数

    戻り値：
      (退職一時金額面, 税額, 手取り額)
    """
    if amount is None:
        amount = 0

    if service_years is None:
        service_years = 0

    retirement_deduction = calc_retirement_income_deduction(service_years)

    taxable_retirement_income = max(
        (amount - retirement_deduction) / 2,
        0,
    )

    income_tax = calc_retirement_income_tax(taxable_retirement_income)

    # 簡易版：住民税は課税退職所得金額の10％として計算
    resident_tax = round(taxable_retirement_income * 0.10)

    total_tax = income_tax + resident_tax

    net_amount = amount - total_tax

    return amount, total_tax, net_amount

# =========================================================
# 退職一時金の入力チェック
# 一時金の場合、勤続年数が未入力なら試算表を止める
# =========================================================
def validate_retirement_lump_sum_input(rows, prefix, person_label):
    """
    prefix:
      あなた   → "you"
      配偶者   → "spouse"

    person_label:
      表示用の名前
    """
    pension_kind = get_text_by_key(rows, f"{prefix}_pension_kind")
    start_age = get_number_by_key(rows, f"{prefix}_start_age")
    amount = get_number_by_key(rows, f"{prefix}_amount")
    service_years = get_number_by_key(rows, f"{prefix}_service_years")

    pension_kind_text = str(pension_kind or "").strip()

    if "一時" not in pension_kind_text:
        return False

    if start_age is None:
        return False

    if amount is None or amount == 0:
        return False

    if service_years is None:
        st.warning(
            f"{person_label}の退職一時金は「一時金」が選択されています。"
            "退職一時金の税引き後手取り額を計算するため、勤続年数を入力してください。"
        )
        return True

    return False

# =========================================================
# 退職一時金・企業年金を年齢に応じて取得する
# 一時金は税引き後手取り、企業年金は額面で返す
# =========================================================
def get_company_pension_income(rows, prefix, age):
    """
    prefix:
      あなた   → "you"
      配偶者   → "spouse"

    age:
      判定する年齢

    戻り値：
      (退職一時金額面, 退職一時金税額, 退職一時金手取り, 企業年金)
    """
    lump_sum_gross = 0
    lump_sum_tax = 0
    lump_sum_net = 0
    pension_income = 0

    pension_kind = get_text_by_key(rows, f"{prefix}_pension_kind")
    service_years = get_number_by_key(rows, f"{prefix}_service_years")
    start_age = get_number_by_key(rows, f"{prefix}_start_age")
    period = get_number_by_key(rows, f"{prefix}_period")
    amount = get_number_by_key(rows, f"{prefix}_amount")

    if start_age is None or amount is None:
        return lump_sum_gross, lump_sum_tax, lump_sum_net, pension_income

    pension_kind_text = str(pension_kind or "").strip()

    # 一時金の場合：開始年齢に1回だけ受け取る
    if "一時" in pension_kind_text:
        if age == start_age:
            lump_sum_gross, lump_sum_tax, lump_sum_net = calc_retirement_lump_sum_net(
                amount,
                service_years,
            )

        return lump_sum_gross, lump_sum_tax, lump_sum_net, pension_income

    # 終身の場合：開始年齢以降ずっと受け取る
    if "終身" in pension_kind_text:
        if age >= start_age:
            pension_income = amount

        return lump_sum_gross, lump_sum_tax, lump_sum_net, pension_income

    # 有期の場合：開始年齢から period 年間受け取る
    if "有期" in pension_kind_text:
        if period is None:
            return lump_sum_gross, lump_sum_tax, lump_sum_net, pension_income

        end_age = start_age + period - 1

        if start_age <= age <= end_age:
            pension_income = amount

        return lump_sum_gross, lump_sum_tax, lump_sum_net, pension_income

    return lump_sum_gross, lump_sum_tax, lump_sum_net, pension_income


# =========================================================
# 公的年金を年齢に応じて取得する
# 基礎年金・厚生年金を別々に返す
# =========================================================
def get_public_pension_income(rows, prefix, age):
    """
    prefix:
      あなた   → "you"
      配偶者   → "spouse"

    age:
      判定する年齢

    戻り値：
      (基礎年金, 厚生年金)
    """
    basic_income = 0
    employee_income = 0

    # 老齢基礎年金
    basic_pension = get_number_by_key(rows, f"{prefix}_basic_pension")
    basic_start_age = get_number_by_key(rows, f"{prefix}_basic_pension_start_age")

    if basic_pension is not None and basic_start_age is not None:
        if age >= basic_start_age:
            basic_income = basic_pension

    # 老齢厚生年金
    employee_pension = get_number_by_key(rows, f"{prefix}_employee_pension")
    employee_start_age = get_number_by_key(rows, f"{prefix}_employee_pension_start_age")

    if employee_pension is not None and employee_start_age is not None:
        if age >= employee_start_age:
            employee_income = employee_pension

    return basic_income, employee_income

# =========================================================
# 年金控除額を計算する
# 公的年金＋企業年金に控除率を掛ける
# =========================================================
def calc_pension_deduction(pension_total, rate):
    """
    pension_total:
      公的年金＋企業年金の合計額（万円）

    rate:
      控除率（％）
      例：15 → 15％控除

    戻り値：
      控除額（万円）
    """
    if pension_total is None:
        pension_total = 0

    if rate is None:
        rate = 0

    return round(pension_total * rate / 100)

# =========================================================
# 雇用保険を年齢に応じて取得する
# =========================================================
def get_unemployment_insurance_income(rows, prefix, age):
    """
    prefix:
      あなた   → "you"
      配偶者   → "spouse"

    age:
      判定する年齢

    戻り値：
      指定年齢に受け取る雇用保険額（万円）
    """
    receive_age = get_number_by_key(rows, f"{prefix}_unemployment_insurance_age")
    amount = get_number_by_key(rows, f"{prefix}_unemployment_insurance_amount")

    if receive_age is None or amount is None:
        return 0

    if age == receive_age:
        return amount

    return 0

# =========================================================
# 個人年金を年齢に応じて取得する
# =========================================================
def get_personal_pension_income(rows, prefix, age):
    """
    prefix:
      あなた   → "you"
      配偶者   → "spouse"

    age:
      判定する年齢

    戻り値：
      指定年齢に受け取る個人年金額（万円）
    """
    start_age = get_number_by_key(rows, f"{prefix}_personal_pension_start_age")
    period = get_number_by_key(rows, f"{prefix}_personal_pension_start_period")
    amount = get_number_by_key(rows, f"{prefix}_personal_pension_start_amount")

    if start_age is None or period is None or amount is None:
        return 0

    end_age = start_age + period - 1

    if start_age <= age <= end_age:
        return amount

    return 0

# =========================================================
# その他収入を年齢に応じて取得する
# =========================================================
def get_other_income(rows, age, periods=2):
    """
    age:
      判定する年齢

    periods:
      その他収入の入力行数
    """
    total = 0

    for i in range(1, periods + 1):
        start_age = get_number_by_key(rows, f"other_income{i}_start_age")
        period = get_number_by_key(rows, f"other_income{i}_period")
        amount = get_number_by_key(rows, f"other_income{i}_amount")

        if start_age is None or period is None or amount is None:
            continue

        end_age = start_age + period - 1

        if start_age <= age <= end_age:
            total += amount

    return total


# =========================================================
# 年齢に応じた期間支出を取得する
# 例：基礎生活費、教育費、住居費、保険料など
# =========================================================
def get_period_expense(rows, base_key, age, periods):
    """
    base_key:
      基礎生活費 → "basic_living_cost"
      教育費     → "educational_cost"
      住居費     → "housing_cost"
      保険料     → "insurance"

    age:
      判定する年齢

    periods:
      期間数
    """
    total = 0

    for i in range(1, periods + 1):
        from_age = get_number_by_key(rows, f"{base_key}{i}_from_age")
        to_age = get_number_by_key(rows, f"{base_key}{i}_to_age")
        amount = get_number_by_key(rows, f"{base_key}{i}_amount")

        if from_age is None or to_age is None or amount is None:
            continue

        if from_age <= age <= to_age:
            total += amount

    return total

# =========================================================
# 開始年齢＋期間で指定する支出を取得する
# 例：その他の支出
# =========================================================
def get_duration_expense(rows, base_key, age, periods):
    """
    base_key:
      その他の支出 → "other_expense"

    age:
      判定する年齢

    periods:
      入力行数
    """
    total = 0

    for i in range(1, periods + 1):
        start_age = get_number_by_key(rows, f"{base_key}{i}_start_age")
        duration = get_number_by_key(rows, f"{base_key}{i}_period")
        amount = get_number_by_key(rows, f"{base_key}{i}_amount")

        if start_age is None or duration is None or amount is None:
            continue

        end_age = start_age + duration - 1

        if start_age <= age <= end_age:
            total += amount

    return total

# =========================================================
# 一時支出を取得する
# 例：車購入、リフォーム、旅行、子供支援など
# =========================================================
def get_one_time_expense(rows, base_key, age, periods):
    """
    base_key:
      一時支出 → "primary_expense"

    age:
      判定する年齢

    periods:
      入力行数
    """
    total = 0

    for i in range(1, periods + 1):
        expense_age = get_number_by_key(rows, f"{base_key}{i}_age")
        amount = get_number_by_key(rows, f"{base_key}{i}_amount")

        if expense_age is None or amount is None:
            continue

        if age == expense_age:
            total += amount

    return total

# =========================================================
# 見出しの横に 表示/非表示 ボタンを置く
# =========================================================
def render_subheader_with_toggle(title, toggle_key):
    """
    title:
      見出し名。例：基礎生活費

    toggle_key:
      表示/非表示状態を保存するためのキー。
      例："show_basic_living_cost"
    """
    if toggle_key not in st.session_state:
        st.session_state[toggle_key] = False

    button_label = "閉じる ▲" if st.session_state[toggle_key] else "表示 ▼"

    title_col, button_col, blank_col = st.columns([1.2, 0.8, 5])

    with title_col:
        st.subheader(title)

    with button_col:
        st.write("")  # 高さ調整
        if st.button(button_label, key=toggle_key + "_button"):
            st.session_state[toggle_key] = not st.session_state[toggle_key]
            st.rerun()

    return st.session_state[toggle_key]


# =========================================================
# 初期値を session_state に入れる
# =========================================================
def init_session_state(rows):
    for item in rows:
        key = str(item["key"]).strip()
        widget_key = f"input_{key}"

        ftype = str(item.get("type") or "text").strip()
        value = item.get("value")

        default_from_key = item.get("default_from_key")

        if (value is None or str(value).strip() == "") and default_from_key not in (None, ""):
            default_value = get_number_by_key(rows, default_from_key)

            default_offset = item.get("default_offset")
            try:
                default_offset = int(default_offset) if default_offset not in (None, "") else 0
            except Exception:
                default_offset = 0

            if default_value is not None:
                value = default_value + default_offset

        display_value = fmt_display(value, ftype)

        if widget_key not in st.session_state:
            st.session_state[widget_key] = display_value

        value_key = f"value_{key}"

        if value_key not in st.session_state:
            st.session_state[value_key] = display_value


# =========================================================
# メイン画面
# =========================================================
st.title("マネープラン入力シート")
st.caption("下記のセルにあなたのデータを入力してください。入力不要な個所は、空欄としてください。")

# Excelを使わず、Python内の INPUT_ITEMS から rows を作る
rows = INPUT_ITEMS
col_map = {}

init_session_state(rows)

tab_current, tab_income, tab_expense, tab_simulation, tab_save = st.tabs(
    ["現況", "収入", "支出", "MP表", "保存"]
)

# =========================================================
# no_excel版：JSONから入力データを読み込む
# 入力欄を表示する前に session_state へ反映する
# =========================================================
with tab_save:
    st.subheader("保存・読込")

    st.info("保存済みのJSONファイルがある場合は、ここで読み込んでください。読み込み後、入力欄に反映されます。")

    uploaded_json = st.file_uploader(
        "保存済みJSONファイルを読み込む",
        type=["json"],
        key="json_uploader",
    )

    if uploaded_json is not None:
        try:
            json_id = f"{uploaded_json.name}_{uploaded_json.size}"

            if st.session_state.get("loaded_json_id") != json_id:
                save_data = json.load(uploaded_json)
                load_save_data_to_session(save_data)

                st.session_state["loaded_json_id"] = json_id
                st.success("JSONファイルを読み込みました。画面に反映します。")
                st.rerun()

        except Exception as e:
            st.error("JSONファイルの読み込みに失敗しました。")
            st.write(e)

# section の表示順を作る
sections = []
for item in rows:
    section = str(item.get("section") or "未分類").strip()
    if section not in sections:
        sections.append(section)


# =========================================================
# 入力欄を表示
# =========================================================

# sectionごとに取り出しやすくする
section_map = {}
for section in sections:
    section_map[section] = [
        item for item in rows
        if str(item.get("section") or "未分類").strip() == section
    ]


def get_section_rows(section_map, section_label):
    """
    section_map から、section_label で始まるセクションを探す。
    """
    if section_label in section_map:
        return section_map[section_label]

    for section_name, section_rows in section_map.items():
        if str(section_name).startswith(section_label):
            return section_rows

    return None

# =========================================================
# 画面デザイン調整CSS
# =========================================================
st.markdown(
    """
    <style>
    /* 通常ボタン */
    div[data-testid="stButton"] button p {
        font-size: 11px !important;
        line-height: 1.0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }

    /* ダウンロードボタン */
    div[data-testid="stDownloadButton"] button p {
        font-size: 11px !important;
        line-height: 1.0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }

    /* 大分類タブ同士の間隔 */
    div[data-testid="stTabs"] [role="tab"] {
        margin-right: 32px !important;
        padding-left: 12px !important;
        padding-right: 12px !important;
    }

    /* 大分類タブの文字 */
    div[data-testid="stTabs"] [role="tab"] p {
        font-size: 24px !important;
        font-weight: 700 !important;
        line-height: 1.2 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────
# 大分類：現在の状況
# ─────────────────────────────
with tab_current:
    st.header("現在の状況")

    # ─────────────────────────────
    # 基本情報
    # ─────────────────────────────
    if "基本情報" in section_map:
        st.subheader("基本情報")
        section_rows = section_map["基本情報"]

        render_row_by_keys(
            section_rows,
            ["basic_year"],
            col_widths=[1, 1, 1, 1, 1, 3],
            all_rows=rows,
        )

        render_row_by_keys(
            section_rows,
            ["you_name", "you_age", "spouse_name", "spouse_age"],
            col_widths=[1, 1, 1, 1, 1, 3],
            all_rows=rows,
        )


    # ─────────────────────────────
    # 子供情報
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "子供情報")

    if section_rows:

        for key_list in make_period_key_rows(
            "child",
            periods=3,
            suffixes=["name", "age"],
        ):
            render_row_by_keys(
                section_rows,
                key_list,
                col_widths=[1, 1, 1, 1, 1, 3],
                all_rows=rows,
            )


    # ─────────────────────────────
    # 金融資産
    # ─────────────────────────────
    if "金融資産" in section_map:
        st.subheader("金融資産")
        section_rows = section_map["金融資産"]

        render_row_by_keys(
            section_rows,
            [
                "cash",
                "time_deposit",
                "insurance",
                "securities",
                "other_financial_assets",
            ],
            col_widths=[1, 1, 1, 1, 1, 3],
            all_rows=rows,
        )


    # ─────────────────────────────
    # 実物資産
    # ─────────────────────────────
    if "実物資産" in section_map:
        st.subheader("実物資産")
        section_rows = section_map["実物資産"]

        render_row_by_keys(
            section_rows,
            [
                "land_house",
                "car",
                "other_fixed_assets",
            ],
            col_widths=[1, 1, 1, 1, 1, 3],
            all_rows=rows,
        )


    # ─────────────────────────────
    # 負債
    # ─────────────────────────────
    if "負債" in section_map:
        st.subheader("負債")
        section_rows = section_map["負債"]

        render_row_by_keys(
            section_rows,
            [
                "credit_card",
                "car_loan",
                "home_loan",
                "other_debts",
            ],
            col_widths=[1, 1, 1, 1, 1, 3],
            all_rows=rows,
        )


# ─────────────────────────────
# 大分類：収入
# ─────────────────────────────
with tab_income:
    st.header("収入")

    # ─────────────────────────────
    # 給与・賞与
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "給与・賞与")

    if section_rows:
               
        show_more = render_subheader_with_toggle(
            "年収/あなた",
            toggle_key="show_you_salary",
        )

        render_rows_with_more(
            section_rows,
            make_period_key_rows(
                "you_salary_",
                periods=4,
                suffixes=["from_age", "to_age", "amount", "takehome_rate"],
            ),
            col_widths=[1, 1, 1, 1, 1, 4],
            all_rows=rows,
            show_more=show_more,
        )
        
        show_more = render_subheader_with_toggle(
            "年収/配偶者",
            toggle_key="show_spouse_salary",
        )

        render_rows_with_more(
            section_rows,
            make_period_key_rows(
                "spouse_salary_",
                periods=4,
                suffixes=["from_age", "to_age", "amount", "takehome_rate"],
            ),
            col_widths=[1, 1, 1, 1, 1, 4],
            all_rows=rows,
            show_more=show_more,
        )


    # ─────────────────────────────
    # 退職一時金・企業年金
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "退職一時金・企業年金")

    if section_rows:
        st.divider()
        st.subheader("退職一時金・企業年金")

        st.markdown("**あなた**")

        render_row_by_keys(
            section_rows,
            [
                "you_pension_kind",
                "you_service_years",
                "you_start_age",
                "you_period",
                "you_amount",
            ],
            col_widths=[1, 1, 1, 1, 1, 4],
            all_rows=rows,
        )      

        st.markdown("配偶者")

        render_row_by_keys(
            section_rows,
            [
                "spouse_pension_kind",
                "spouse_service_years",
                "spouse_start_age",
                "spouse_period",
                "spouse_amount",
            ],
            col_widths=[1, 1, 1, 1, 1, 4],
            all_rows=rows,
        )


    # ─────────────────────────────
    # 公的年金
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "公的年金")

    if section_rows:
        st.divider()
        st.subheader("公的年金")

        st.markdown("**あなた**")

        render_row_by_keys(
            section_rows,
            [
                "you_basic_pension",
                "you_basic_pension_start_age",
                "you_employee_pension",
                "you_employee_pension_start_age",
            ],
            col_widths=[1.2, 0.8, 1.2, 0.8, 1, 4],
            all_rows=rows,
        )

        st.markdown("配偶者")

        render_row_by_keys(
            section_rows,
            [
                "spouse_basic_pension",
                "spouse_basic_pension_start_age",
                "spouse_employee_pension",
                "spouse_employee_pension_start_age",
            ],
            col_widths=[1.2, 0.8, 1.2, 0.8, 1, 4],
            all_rows=rows,
        )

        st.markdown("**控除率**")

        render_row_by_keys(
            section_rows,
            [
                "public_pension_rate",
            ],
            col_widths=[1.2, 1, 1, 1, 1, 4],
            all_rows=rows,
        )


    # ─────────────────────────────
    # その他の収入
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "その他の収入")

    if section_rows:
        st.divider()
        st.subheader("その他の収入")

        st.markdown("**雇用保険**")

        render_row_by_keys(
            section_rows,
            [
                "you_unemployment_insurance_age",
                "you_unemployment_insurance_amount",
                "spouse_unemployment_insurance_age",
                "spouse_unemployment_insurance_amount",
            ],
            col_widths=[1.2, 0.8, 1.2, 0.8, 1, 4],
            all_rows=rows,
        )

        st.markdown("**個人年金**")

        render_row_by_keys(
            section_rows,
            [
                "you_personal_pension_start_age",
                "you_personal_pension_start_period",
                "you_personal_pension_start_amount",
            ],
            col_widths=[1, 1, 1, 1, 1, 4],
            all_rows=rows,
        )

        render_row_by_keys(
            section_rows,
            [
                "spouse_personal_pension_start_age",
                "spouse_personal_pension_start_period",
                "spouse_personal_pension_start_amount",
            ],
            col_widths=[1, 1, 1, 1, 1, 4],
            all_rows=rows,
        )

        st.markdown("**その他収入**")

        render_row_by_keys(
            section_rows,
            [
                "other_income1_start_age",
                "other_income1_period",
                "other_income1_amount",
            ],
            col_widths=[1, 1, 1, 1, 1, 4],
            all_rows=rows,
        )

        render_row_by_keys(
            section_rows,
            [
                "other_income2_start_age",
                "other_income2_period",
                "other_income2_amount",
            ],
            col_widths=[1, 1, 1, 1, 1, 4],
            all_rows=rows,
        )


# ─────────────────────────────
# 大分類：支出
# ─────────────────────────────
with tab_expense:
    st.header("支出")

    # ─────────────────────────────
    # 基礎生活費
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "基礎生活費")

    if section_rows:
        show_more = render_subheader_with_toggle(
            "基礎生活費",
            toggle_key="show_basic_living_cost",
        )

        render_rows_with_more(
            section_rows,
            make_period_key_rows(
                "basic_living_cost",
                periods=5,
                suffixes=["from_age", "to_age", "amount"],
            ),
            col_widths=[1.4, 1.4, 1.4, 2, 2, 2],
            all_rows=rows,
            show_more=show_more,
        )


    # ─────────────────────────────
    # 教育費
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "教育費")

    if section_rows:
        st.divider()
        show_more = render_subheader_with_toggle(
        "教育費",
        toggle_key="show_educational_cost",
        )

        render_rows_with_more(
            section_rows,
            make_period_key_rows(
                "educational_cost",
                periods=4,
                suffixes=["from_age", "to_age", "amount"],
            ),
            col_widths=[1.4, 1.4, 1.4, 2, 2, 2],
            all_rows=rows,
            show_more=show_more,
        )


    # ─────────────────────────────
    # 住居費
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "住居費")

    if section_rows:
        st.divider()
        show_more = render_subheader_with_toggle(
        "住居費",
        toggle_key="show_housing_cost",
        )

        render_rows_with_more(
            section_rows,
            make_period_key_rows(
                "housing_cost",
                periods=4,
                suffixes=["from_age", "to_age", "amount"],
            ),
            col_widths=[1.4, 1.4, 1.4, 2, 2, 2],
            all_rows=rows,
            show_more=show_more,
        )


    # ─────────────────────────────
    # 保険料
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "保険料")

    if section_rows:
        st.divider()
        show_more = render_subheader_with_toggle(
        "保険料",
        toggle_key="show_insurance",
        )

        render_rows_with_more(
            section_rows,
            make_period_key_rows(
                "insurance",
                periods=4,
                suffixes=["from_age", "to_age", "amount"],
            ),
            col_widths=[1.4, 1.4, 1.4, 2, 2, 2],
            all_rows=rows,
            show_more=show_more,
        )


    # ─────────────────────────────
    # その他の支出
    # ─────────────────────────────
    section_rows = get_section_rows(section_map, "その他の支出")

    if section_rows:
        st.divider()
        
        show_more = render_subheader_with_toggle(
        "その他の支出",
        toggle_key="show_other_expense",
        )

        render_rows_with_more(
            section_rows,
            make_period_key_rows(
                "other_expense",
                periods=4,
                suffixes=["start_age", "period", "amount"],
            ),
            col_widths=[1.4, 1.4, 1.4, 2, 2, 2],
            all_rows=rows,
            show_more=show_more,
        )

        show_more = render_subheader_with_toggle(
        "一時支出",
        toggle_key="show_primary_expense",
        )

        render_rows_with_more(
            section_rows,
            make_period_key_rows(
                "primary_expense",
                periods=10,
                suffixes=["name", "age", "amount"],
            ),
            col_widths=[1.4, 1.4, 1.4, 2, 2, 2],
            all_rows=rows,
            show_more=show_more,
        )


    def get_section_rows(section_map, section_label):
        """
        section_map から、section_label で始まるセクションを探す。
        例：
          section_label = "教育費"
          Excel側 section = "教育費キョウイクヒ"
        でも拾えるようにする。
        """
        # 完全一致を優先
        if section_label in section_map:
            return section_map[section_label]

        # 前方一致で探す
        for section_name, section_rows in section_map.items():
            if str(section_name).startswith(section_label):
                return section_rows

        return None


    # =========================================================
    # 保存してダウンロード
    # =========================================================
    # st.divider()
    #
    # if st.button("①入力データをエクセル化する"):
    #     wb = load_workbook(INPUT_FILE)
    #     ws = wb[SHEET_NAME]
    #
    #     value_col = col_map["value"]
    #
    #    for item in rows:
    #        key = str(item["key"]).strip()
    #        widget_key = f"input_{key}"
    #
    #        ftype = str(item.get("type") or "text").strip()
    #        row_no = item["_row_no"]
    #
    #        screen_value = st.session_state.get(widget_key, "")
    #        excel_value = to_excel_value(screen_value, ftype)
    #
    #        ws.cell(row=row_no, column=value_col).value = excel_value
    #
    #    output = BytesIO()
    #    wb.save(output)
    #    output.seek(0)
    #
    #    st.session_state["download_excel"] = output.getvalue()
    #
    #if "download_excel" in st.session_state:
    #    st.download_button(
    #        label="②Excelをダウンロード",
    #       data=st.session_state["download_excel"],
    #       file_name="mp_edited.xlsx",
    #       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #    )
    
# =========================================================
# no_excel版：入力データをJSONで保存
# =========================================================
with tab_save:
    st.divider()

    st.info("現在の入力内容をJSONファイルとして保存できます。")

    save_data = create_save_data(rows)

    json_text = json.dumps(
        save_data,
        ensure_ascii=False,
        indent=2,
    )

    st.download_button(
        label="入力データを保存（JSON）",
        data=json_text,
        file_name="money_plan_data.json",
        mime="application/json",
    )

# ─────────────────────────────
# 大分類：試算表
# ─────────────────────────────
with tab_simulation:
    st.header("マネープラン表")

    st.info("まずは、金融資産残高を毎年繰り越す形のマネープラン表を作成します。")

    # 基本情報
    basic_year = get_number_by_key(rows, "basic_year") or 2025
    you_age = get_number_by_key(rows, "you_age") or 65
    spouse_age = get_number_by_key(rows, "spouse_age") or 65

    # 退職一時金の入力チェック
    has_retirement_error = False

    has_retirement_error = validate_retirement_lump_sum_input(
        rows,
        "you",
        "あなた",
    ) or has_retirement_error

    has_retirement_error = validate_retirement_lump_sum_input(
        rows,
        "spouse",
        "配偶者",
    ) or has_retirement_error

    if has_retirement_error:
        st.stop()

    # 金融資産合計
    initial_financial_assets = 0
    for key in [
        "cash",
        "time_deposit",
        "insurance",
        "securities",
        "other_financial_assets",
    ]:
        initial_financial_assets += get_number_by_key(rows, key) or 0

    # 試算期間：あなたの年齢から100歳まで
    plan_rows = []

    financial_assets = initial_financial_assets

    for age in range(you_age, 101):
        year = basic_year + (age - you_age)
        spouse_current_age = spouse_age + (age - you_age)

        # 収入
        you_salary_income = get_salary_income(rows, "you", age, periods=4)
        spouse_salary_income = get_salary_income(rows, "spouse", spouse_current_age, periods=4)

        you_basic_pension, you_employee_pension = get_public_pension_income(
            rows,
            "you",
            age,
        )

        spouse_basic_pension, spouse_employee_pension = get_public_pension_income(
            rows,
            "spouse",
            spouse_current_age,
        )
        you_retirement_lump_sum_gross, you_retirement_lump_sum_tax, you_retirement_lump_sum_net, you_company_pension = get_company_pension_income(
            rows,
            "you",
            age,
        )

        spouse_retirement_lump_sum_gross, spouse_retirement_lump_sum_tax, spouse_retirement_lump_sum_net, spouse_company_pension = get_company_pension_income(
            rows,
            "spouse",
            spouse_current_age,
        )

        # 年金控除率
        public_pension_rate = get_number_by_key(rows, "public_pension_rate") or 0

        # 控除対象となる年金収入
        # 公的年金 ＋ 企業年金
        pension_income_total = (
            you_company_pension
            + spouse_company_pension
            + you_basic_pension
            + you_employee_pension
            + spouse_basic_pension
            + spouse_employee_pension
        )

        # 年金控除額
        pension_deduction = calc_pension_deduction(
            pension_income_total,
            public_pension_rate,
        )

        # 表示用はマイナスにする
        pension_deduction_display = -pension_deduction

        you_unemployment_income = get_unemployment_insurance_income(
            rows,
            "you",
            age,
        )

        spouse_unemployment_income = get_unemployment_insurance_income(
            rows,
            "spouse",
            spouse_current_age,
        )

        you_personal_pension = get_personal_pension_income(
            rows,
            "you",
            age,
        )

        spouse_personal_pension = get_personal_pension_income(
            rows,
            "spouse",
            spouse_current_age,
        )

        other_income = get_other_income(
            rows,
            age=age,
            periods=2,
        )


        # 収入合計
        income_total = (
            you_salary_income
            + spouse_salary_income
            + you_retirement_lump_sum_net
            + spouse_retirement_lump_sum_net
            + you_company_pension
            + spouse_company_pension
            + you_basic_pension
            + you_employee_pension
            + spouse_basic_pension
            + spouse_employee_pension
            + you_unemployment_income
            + spouse_unemployment_income
            + you_personal_pension
            + spouse_personal_pension
            + other_income
            - pension_deduction
        )

        # 支出
        basic_living_cost = get_period_expense(
            rows,
            base_key="basic_living_cost",
            age=age,
            periods=5,
        )

        educational_cost = get_period_expense(
            rows,
            base_key="educational_cost",
            age=age,
            periods=4,
        )

        housing_cost = get_period_expense(
            rows,
            base_key="housing_cost",
            age=age,
            periods=4,
        )

        insurance_cost = get_period_expense(
            rows,
            base_key="insurance",
            age=age,
            periods=4,
        )

        other_expense = get_duration_expense(
            rows,
            base_key="other_expense",
            age=age,
            periods=4,
        )

        one_time_expense = get_one_time_expense(
            rows,
            base_key="primary_expense",
            age=age,
            periods=10,
        )

        expense_total = (
            basic_living_cost
            + educational_cost
            + housing_cost
            + insurance_cost
            + other_expense
            + one_time_expense
        )

        annual_balance = income_total - expense_total

        # 金融資産残高を更新
        financial_assets = financial_assets + annual_balance

        plan_rows.append({
            "年": year,
            "あなたの年齢": age,
            "配偶者の年齢": spouse_current_age,
            "あなたの給与（万円）": you_salary_income,
            "配偶者の給与（万円）": spouse_salary_income,
            "あなたの退職一時金（万円）": you_retirement_lump_sum_net,
            "配偶者の退職一時金（万円）": spouse_retirement_lump_sum_net,
            "あなたの企業年金（万円）": you_company_pension,
            "配偶者の企業年金（万円）": spouse_company_pension,
            "あなたの基礎年金（万円）": you_basic_pension,
            "あなたの厚生年金（万円）": you_employee_pension,
            "配偶者の基礎年金（万円）": spouse_basic_pension,
            "配偶者の厚生年金（万円）": spouse_employee_pension,
            "年金控除額（万円）": pension_deduction_display,
            "あなたの雇用保険（万円）": you_unemployment_income,
            "配偶者の雇用保険（万円）": spouse_unemployment_income,
            "あなたの個人年金（万円）": you_personal_pension,
            "配偶者の個人年金（万円）": spouse_personal_pension,
            "その他収入（万円）": other_income,
            "収入合計（万円）": income_total,
            "基礎生活費（万円）": basic_living_cost,
            "教育費（万円）": educational_cost,
            "住居費（万円）": housing_cost,
            "保険料（万円）": insurance_cost,
            "その他の支出（万円）": other_expense,
            "一時支出（万円）": one_time_expense,
            "支出合計（万円）": expense_total,
            "年間収支（万円）": annual_balance,
            "金融資産残高（万円）": financial_assets,
        })

    df_plan = pd.DataFrame(plan_rows)

    # 表示用に縦横を入れ替える
    df_display = df_plan.T

    # 横方向の列見出しに「西暦」と「あなたの年齢」を入れる
    df_display.columns = [
        f"{year}年\n{age}歳"
        for year, age in zip(df_plan["年"], df_plan["あなたの年齢"])
    ]

    # 本体側から「年」と「あなたの年齢」は外す
    # その代わり、列見出しに表示する
    df_display = df_display.drop(
        index=["年", "あなたの年齢"],
        errors="ignore",
    )

    # 左上の「項目」を表示
    df_display.index.name = "西暦・あなたの年齢"
    
    st.subheader("簡易試算表")
    st.dataframe(
    df_display,
    use_container_width=True,
    )  

# =========================================================
# 試算表の行ごとに背景色を付ける
# =========================================================
def style_plan_rows(row):
    item_name = str(row.name)

    # 収入系：薄い暖色
    income_keywords = [
        "給与",
        "退職一時金",
        "企業年金",
        "基礎年金",
        "厚生年金",
        "雇用保険",
        "個人年金",
        "その他収入",
        "収入合計",
    ]

    # 支出系：薄い寒色
    expense_keywords = [
        "基礎生活費",
        "教育費",
        "住居費",
        "保険料",
        "その他の支出",
        "一時支出",
        "支出合計",
    ]

    # 集計系：薄いグレー
    summary_keywords = [
        "年間収支",
        "金融資産残高",
    ]

    # 控除系：薄い黄色
    deduction_keywords = [
        "控除",
        "税額",
    ]

    if any(keyword in item_name for keyword in income_keywords):
        return ["background-color: #fff2cc"] * len(row)

    if any(keyword in item_name for keyword in expense_keywords):
        return ["background-color: #ddebf7"] * len(row)

    if any(keyword in item_name for keyword in deduction_keywords):
        return ["background-color: #fce4d6"] * len(row)

    if any(keyword in item_name for keyword in summary_keywords):
        return ["background-color: #e7e6e6; font-weight: bold"] * len(row)

    return [""] * len(row)

    